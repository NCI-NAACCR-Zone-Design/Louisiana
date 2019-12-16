#!/bin/env python3

import sys
import csv
import openpyxl

import settings


class IncidenceDatasetConverter:
    def run(self):
        print("IncidenceDatasetConverter")

        print("    Reading {}".format(settings.INPUT_CANCERXLS))
        wb = openpyxl.load_workbook(filename=settings.INPUT_CANCERXLS, read_only=True)
        ws = wb[settings.INPUT_CANCERXLS_SHEETNAME]

        try:
            assert ws['A1'].value == 'ID'
            assert ws['B1'].value == 'sex'
            assert ws['C1'].value == 'cancer'
            assert ws['D1'].value == 'years'
            assert ws['E1'].value == 'Zones'
            assert ws['F1'].value == 'W_PopTot'
            assert ws['G1'].value == 'W_Cases'
            assert ws['H1'].value == 'W_AAIR'
            assert ws['I1'].value == 'W_LCI'
            assert ws['J1'].value == 'W_UCI'
            assert ws['K1'].value == 'B_PopTot'
            assert ws['L1'].value == 'B_Cases'
            assert ws['M1'].value == 'B_AAIR'
            assert ws['N1'].value == 'B_LCI'
            assert ws['O1'].value == 'B_UCI'
            assert ws['P1'].value == 'H_PopTot'
            assert ws['Q1'].value == 'H_Cases'
            assert ws['R1'].value == 'H_AAIR'
            assert ws['S1'].value == 'H_LCI'
            assert ws['T1'].value == 'H_UCI'
            assert ws['U1'].value == 'A_PopTot'
            assert ws['V1'].value == 'A_Cases'
            assert ws['W1'].value == 'A_AAIR'
            assert ws['X1'].value == 'A_LCI'
            assert ws['Y1'].value == 'A_UCI'
            assert ws['Z1'].value == 'PopTot'
            assert ws['AA1'].value == 'Cases'
            assert ws['AB1'].value == 'AAIR'
            assert ws['AC1'].value == 'LCI'
            assert ws['AD1'].value == 'UCI'
        except AssertionError:
            print("    ERROR: Headings in first row are not as expected")
            sys.exit(10)

        with open(settings.OUTPUT_INCIDENCECSV, 'w') as csvfile:
            csvwriter = csv.writer(csvfile)

            # header row
            csvwriter.writerow([
                "Zone",
                "sex", "cancer", "years",
                "PopTot", "Cases", "AAIR", "LCI", "UCI",
                "W_PopTot", "W_Cases", "W_AAIR", "W_LCI", "W_UCI",
                "B_PopTot", "B_Cases", "B_AAIR", "B_LCI", "B_UCI",
                "H_PopTot", "H_Cases", "H_AAIR", "H_LCI", "H_UCI",
                "A_PopTot", "A_Cases", "A_AAIR", "A_LCI", "A_UCI",
            ])

            # data rows
            for row in ws.iter_rows(min_row=2):
                if not row[0].value:  # skip blank rows
                    continue
                csvwriter.writerow([
                    row[4].value,
                    row[1].value, row[2].value, row[3].value,
                    row[25].value, row[26].value, row[27].value, row[28].value, row[29].value,
                    row[5].value, row[6].value, row[7].value, row[8].value, row[9].value,
                    row[10].value, row[11].value, row[12].value, row[13].value, row[14].value,
                    row[15].value, row[16].value, row[17].value, row[18].value, row[19].value,
                    row[20].value, row[21].value, row[22].value, row[23].value, row[24].value,
                ])

        print("    Wrote {}".format(settings.OUTPUT_INCIDENCECSV))


if __name__ == '__main__':
    IncidenceDatasetConverter().run()
    print("DONE")
