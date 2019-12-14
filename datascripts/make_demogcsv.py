#!/bin/env python3

import sys
import csv
import openpyxl

import settings


class DemographicDatasetConverter:
    def run(self):
        print("DemographicDatasetConverter")

        print("    Reading {}".format(settings.INPUT_DEMOGSXLS))
        wb = openpyxl.load_workbook(filename=settings.INPUT_DEMOGSXLS, read_only=True)
        ws = wb[settings.INPUT_DEMOGXLS_SHEETNAME]

        try:
            assert ws['A1'].value == 'Zone'
            assert ws['B1'].value == 'QNSES'
            assert ws['C1'].value == 'PerRural'
            assert ws['D1'].value == 'PerUninsured'
            assert ws['E1'].value == 'PerForeignBorn'
            assert ws['F1'].value == 'PerWhite'
            assert ws['G1'].value == 'PerBlack'
            assert ws['H1'].value == 'PerAPI'
            assert ws['I1'].value == 'PerHispanic'
            assert ws['J1'].value == 'PopAll'
        except AssertionError:
            print("    ERROR: Headings in first row are not as expected")
            sys.exit(10)

        with open(settings.OUTPUT_DEMOGCSV, 'w') as csvfile:
            csvwriter = csv.writer(csvfile)

            # header row
            csvwriter.writerow([
                "Zone",
                "QNSES",
                "PerRural",
                "PerUninsured",
                "PerForeignBorn",
                "PerWhite",
                "PerBlack",
                "PerAPI",
                "PerHispanic",
                "PopAll",
            ])

            # data rows
            for row in ws.iter_rows(min_row=2):
                if not len(row):  # skip blank rows, often present at the end
                    continue
                csvwriter.writerow([
                    row[0].value,
                    row[1].value,
                    row[2].value,
                    row[3].value,
                    row[4].value,
                    row[5].value,
                    row[6].value,
                    row[7].value,
                    row[8].value,
                    row[9].value,
                ])

        print("    Wrote {}".format(settings.OUTPUT_DEMOGCSV))


if __name__ == '__main__':
    DemographicDatasetConverter().run()
    print("DONE")
